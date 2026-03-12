"""
问卷交叉分析 MCP Server - Excel 导出模块 (v2)

专业级可视化：清晰配色、斑马纹、总计行高亮、DataBar、冻结窗格
"""

import os
import json
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.formatting.rule import DataBarRule


# ========================================================================= #
#                         配色方案 (专业蓝灰系)
# ========================================================================= #

class Theme:
    """统一配色主题"""
    # 标题行
    HEADER_BG = "2F5496"        # 深蓝
    HEADER_FONT = "FFFFFF"      # 白色字
    # 索引列（问题 / 选项）
    INDEX_BG = "D6E4F0"         # 浅蓝
    INDEX_FONT = "1F3864"       # 深蓝字
    # 斑马纹
    ROW_EVEN_BG = "F2F2F2"      # 浅灰（偶数行）
    ROW_ODD_BG = "FFFFFF"       # 白色（奇数行）
    # 总计行
    TOTAL_BG = "E2EFDA"         # 浅绿
    TOTAL_FONT = "375623"       # 深绿字
    # DataBar
    FREQ_BAR = "5B9BD5"         # 蓝色数据条
    PCT_BAR = "ED7D31"          # 橙色数据条
    # 边框
    BORDER_COLOR = "B4C6E7"     # 浅蓝边框
    # 得分
    SCORE_HEADER_BG = "4472C4"  # 得分表头蓝
    # 报告
    REPORT_HEADER_BG = "2F5496"
    REPORT_FINDING_BG = "FFF2CC"  # 浅黄（关键发现列底色）
    # 字体
    FONT_NAME = "微软雅黑"
    FONT_NAME_FALLBACK = "Arial"


def _thin_border():
    side = Side(style='thin', color=Theme.BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill(start_color=Theme.HEADER_BG, end_color=Theme.HEADER_BG, fill_type="solid")


def _header_font(size=11):
    return Font(name=Theme.FONT_NAME, size=size, bold=True, color=Theme.HEADER_FONT)


def _index_fill():
    return PatternFill(start_color=Theme.INDEX_BG, end_color=Theme.INDEX_BG, fill_type="solid")


def _index_font(bold=False):
    return Font(name=Theme.FONT_NAME, size=10, bold=bold, color=Theme.INDEX_FONT)


def _total_fill():
    return PatternFill(start_color=Theme.TOTAL_BG, end_color=Theme.TOTAL_BG, fill_type="solid")


def _total_font():
    return Font(name=Theme.FONT_NAME, size=10, bold=True, color=Theme.TOTAL_FONT)


def _body_font():
    return Font(name=Theme.FONT_NAME, size=10)


def _even_fill():
    return PatternFill(start_color=Theme.ROW_EVEN_BG, end_color=Theme.ROW_EVEN_BG, fill_type="solid")


def _odd_fill():
    return PatternFill(start_color=Theme.ROW_ODD_BG, end_color=Theme.ROW_ODD_BG, fill_type="solid")


# ========================================================================= #
#                  格式化交叉分析 / 百分比 sheet
# ========================================================================= #

def _format_crosstab_sheet(ws, is_percent=False):
    """
    格式化交叉分析/百分比 sheet：
    - 标题行：深蓝底+白字
    - 索引列（A, B）：浅蓝底
    - 总计行：浅绿底+加粗
    - 斑马纹
    - DataBar
    - 百分比格式（如适用）
    """
    max_row = ws.max_row
    max_col = ws.max_column
    border = _thin_border()
    center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
    left_align = Alignment(wrap_text=True, vertical='center', horizontal='left')
    right_align = Alignment(wrap_text=True, vertical='center', horizontal='right')

    # ---- 标题行（第1行）----
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _header_fill()
        cell.font = _header_font(size=10)
        cell.alignment = center_align
        cell.border = border

    # ---- 找出总计行 ----
    total_rows = set()
    for row_idx in range(2, max_row + 1):
        # 检查 B 列（选项列）是否为"总计"
        b_val = ws.cell(row=row_idx, column=2).value
        if b_val and str(b_val).strip() in ("总计", "合计", "Total"):
            total_rows.add(row_idx)

    # ---- 数据行 ----
    data_row_count = 0
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 22
        is_total = row_idx in total_rows
        data_row_count += 1

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if is_total:
                # 总计行
                cell.fill = _total_fill()
                cell.font = _total_font()
                if col_idx <= 2:
                    cell.alignment = left_align
                else:
                    cell.alignment = right_align
                    if is_percent:
                        cell.number_format = '0%'
            elif col_idx <= 2:
                # 索引列（问题 / 选项）
                cell.fill = _index_fill()
                cell.font = _index_font(bold=(col_idx == 1))
                cell.alignment = left_align
            else:
                # 数据列 — 斑马纹
                if data_row_count % 2 == 0:
                    cell.fill = _even_fill()
                else:
                    cell.fill = _odd_fill()
                cell.font = _body_font()
                cell.alignment = right_align

                # 百分比格式：显示整数百分比，底层保留完整精度
                if is_percent:
                    cell.number_format = '0%'

    # ---- DataBar ----
    bar_color = Theme.PCT_BAR if is_percent else Theme.FREQ_BAR
    non_total_rows = [r for r in range(2, max_row + 1) if r not in total_rows]

    if non_total_rows:
        for col_idx in range(3, max_col + 1):
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}{min(non_total_rows)}:{col_letter}{max(non_total_rows)}"
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='max',
                color=bar_color,
                showValue="None",
                minLength=0,
                maxLength=100,
            )
            ws.conditional_formatting.add(data_range, rule)

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 30  # 问题
    ws.column_dimensions['B'].width = 22  # 选项
    for col_idx in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ---- 冻结 + 隐藏网格线 ----
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


# ========================================================================= #
#                  格式化得分分析 sheet
# ========================================================================= #

def _format_score_sheet(ws):
    """格式化得分分析 sheet"""
    max_row = ws.max_row
    max_col = ws.max_column
    border = _thin_border()
    center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
    right_align = Alignment(wrap_text=True, vertical='center', horizontal='right')
    left_align = Alignment(wrap_text=True, vertical='center', horizontal='left')

    # 标题行
    score_fill = PatternFill(start_color=Theme.SCORE_HEADER_BG, end_color=Theme.SCORE_HEADER_BG, fill_type="solid")
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = score_fill
        cell.font = _header_font(size=10)
        cell.alignment = center_align
        cell.border = border

    # 数据行
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 28
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx <= 2:
                cell.fill = _index_fill()
                cell.font = _index_font(bold=(col_idx == 1))
                cell.alignment = left_align
            else:
                cell.font = Font(name=Theme.FONT_NAME, size=12, bold=True, color="C00000")
                cell.alignment = center_align
                cell.number_format = '0.00'

    # 列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    for col_idx in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


# ========================================================================= #
#                  格式化分析报告 sheet
# ========================================================================= #

def _format_report_sheet(ws):
    """格式化 AI 分析报告 sheet"""
    max_row = ws.max_row
    max_col = ws.max_column
    border = _thin_border()

    report_header_fill = PatternFill(
        start_color=Theme.REPORT_HEADER_BG,
        end_color=Theme.REPORT_HEADER_BG,
        fill_type="solid",
    )
    finding_fill = PatternFill(
        start_color=Theme.REPORT_FINDING_BG,
        end_color=Theme.REPORT_FINDING_BG,
        fill_type="solid",
    )

    # 标题行
    ws.row_dimensions[1].height = 35
    for cell in ws[1]:
        cell.fill = report_header_fill
        cell.font = _header_font(size=12)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = border

    # 数据行
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            if col_idx == 1:
                # 题目列
                cell.font = Font(name=Theme.FONT_NAME, size=11, bold=True, color=Theme.INDEX_FONT)
                cell.fill = _index_fill()
            elif col_idx == 2:
                # 关键发现列
                cell.font = Font(name=Theme.FONT_NAME, size=11)
                cell.fill = finding_fill
            else:
                cell.font = Font(name=Theme.FONT_NAME, size=10, color="666666")
                if row_idx % 2 == 0:
                    cell.fill = _even_fill()

    # 列宽
    ws.column_dimensions['A'].width = 25   # 题目
    ws.column_dimensions['B'].width = 80   # 关键发现
    if max_col >= 3:
        ws.column_dimensions['C'].width = 40  # 详细说明

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ========================================================================= #
#                         导出主函数
# ========================================================================= #

def export_crosstab_excel(
    file_path: str,
    output_path: str,
    report_text: str = "",
    style_overrides: dict = None,
) -> str:
    """
    将缓存的交叉分析结果导出为专业级格式化 Excel。
    """
    from crosstab_engine import _CACHE

    if file_path not in _CACHE or "crosstab_result" not in _CACHE[file_path]:
        raise ValueError("未找到交叉分析结果，请先执行 run_crosstab")

    ct_result = _CACHE[file_path]["crosstab_result"]
    freq_df = ct_result["freq_df"]
    percent_df = ct_result["percent_df"]
    col_labels = ct_result["col_labels"]

    # 删除已有文件
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            raise PermissionError(f"请关闭正在使用的文件：{output_path}")

    score_df = _CACHE[file_path].get("score_df", None)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # ---- Sheet 1: 交叉分析（频数）----
        freq_df.to_excel(writer, sheet_name='交叉分析', merge_cells=True)
        _format_crosstab_sheet(writer.sheets['交叉分析'], is_percent=False)
        writer.sheets['交叉分析'].sheet_properties.tabColor = "B4C6E7"  # 浅蓝

        # ---- Sheet 2: 列百分比 ----
        percent_df.to_excel(writer, sheet_name='列百分比', merge_cells=True)
        _format_crosstab_sheet(writer.sheets['列百分比'], is_percent=True)
        writer.sheets['列百分比'].sheet_properties.tabColor = "F8CBAD"  # 浅橙

        # ---- Sheet 3: 得分分析（如有）----
        if score_df is not None and not score_df.empty:
            score_df.to_excel(writer, sheet_name='得分分析', merge_cells=True)
            _format_score_sheet(writer.sheets['得分分析'])
            writer.sheets['得分分析'].sheet_properties.tabColor = "C6EFCE"  # 浅绿

        # ---- Sheet 4: 分析报告（如有）----
        if report_text:
            _write_report_sheet(writer, report_text)
            writer.sheets['分析报告'].sheet_properties.tabColor = "D9C4EC"  # 浅紫

    return output_path


def _write_report_sheet(writer, report_text: str):
    """写入 AI 分析报告 sheet"""
    try:
        report_data = json.loads(report_text)
        if isinstance(report_data, list):
            report_df = pd.DataFrame(report_data)
            col_mapping = {
                "question": "题目",
                "finding": "关键发现",
                "detail": "详细说明",
                "diff_option": "差异最大选项",
                "diff_value": "最大差异值",
            }
            report_df = report_df.rename(columns={
                k: v for k, v in col_mapping.items() if k in report_df.columns
            })
            report_df.to_excel(writer, sheet_name='分析报告', index=False)
            _format_report_sheet(writer.sheets['分析报告'])
            return
    except (json.JSONDecodeError, TypeError):
        pass

    # 纯文本模式
    ws = writer.book.create_sheet('分析报告')
    ws.cell(row=1, column=1, value="交叉分析报告")
    ws.cell(row=1, column=1).font = Font(name=Theme.FONT_NAME, size=16, bold=True, color=Theme.HEADER_BG)

    lines = report_text.split("\n")
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name=Theme.FONT_NAME, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 120
    ws.sheet_view.showGridLines = False